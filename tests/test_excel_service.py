#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Tests for excel_service.py - ExcelService class
"""

import unittest
import sys
import os
import tempfile
from pathlib import Path

# Add parent directory to path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from tests.conftest import TestFixtures
from excel_service import ExcelService
from models import Host

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


@unittest.skipIf(not OPENPYXL_AVAILABLE, "openpyxl not installed")
class TestExcelService(unittest.TestCase):
    """Tests for ExcelService."""
    
    def setUp(self):
        """Setup temporary directory for Excel files."""
        self.temp_dir = tempfile.mkdtemp()
    
    def tearDown(self):
        """Cleanup temporary files."""
        import shutil
        if os.path.exists(self.temp_dir):
            shutil.rmtree(self.temp_dir)
    
    def create_test_excel(self, filename: str, data: list) -> str:
        """Helper to create test Excel file."""
        filepath = os.path.join(self.temp_dir, filename)
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        
        # Headers
        sheet.append(["Название", "IP адрес", "Адрес", "Группа"])
        
        # Data
        for row in data:
            sheet.append(row)
        
        workbook.save(filepath)
        return filepath
    
    def test_export_hosts(self):
        """Test exporting hosts to Excel."""
        hosts = TestFixtures.create_sample_hosts(3)
        filepath = os.path.join(self.temp_dir, "export.xlsx")
        
        ExcelService.export_hosts(filepath, hosts)
        
        # Verify file was created
        self.assertTrue(Path(filepath).exists())
        
        # Verify contents
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active
        
        # Check headers
        headers = [cell.value for cell in sheet[1]]
        self.assertIn("Название", headers)
        self.assertIn("IP адрес", headers)
        
        # Check data rows (1 header + 3 data)
        self.assertEqual(sheet.max_row, 4)
    
    def test_export_empty_list(self):
        """Test exporting empty host list."""
        filepath = os.path.join(self.temp_dir, "empty.xlsx")
        
        ExcelService.export_hosts(filepath, [])
        
        # Should still create file with headers
        self.assertTrue(Path(filepath).exists())
        
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active
        
        # Only header row
        self.assertEqual(sheet.max_row, 1)
    
    def test_import_valid_hosts(self):
        """Test importing valid hosts from Excel."""
        data = [
            ["Server1", "192.168.1.1", "Location 1", "Group1"],
            ["Server2", "192.168.1.2", "Location 2", "Group2"],
            ["Server3", "google.com", "Location 3", "Group1"],
        ]
        filepath = self.create_test_excel("import.xlsx", data)
        
        existing_ips = set()
        hosts, skipped, errors = ExcelService.import_hosts(filepath, existing_ips)
        
        self.assertEqual(len(hosts), 3)
        self.assertEqual(skipped, 0)
        self.assertEqual(len(errors), 0)
        
        # Verify data
        self.assertEqual(hosts[0].name, "Server1")
        self.assertEqual(hosts[0].ip, "192.168.1.1")
        self.assertEqual(hosts[1].name, "Server2")
        self.assertEqual(hosts[2].ip, "google.com")
    
    def test_import_with_duplicates(self):
        """Test importing hosts with duplicate IPs."""
        data = [
            ["Server1", "192.168.1.1", "Location 1", "Group1"],
            ["Server2", "192.168.1.2", "Location 2", "Group2"],
        ]
        filepath = self.create_test_excel("import.xlsx", data)
        
        # Existing IPs set
        existing_ips = {"192.168.1.1"}
        
        hosts, skipped, errors = ExcelService.import_hosts(filepath, existing_ips)
        
        # Should skip duplicate
        self.assertEqual(len(hosts), 1)
        self.assertEqual(skipped, 1)
        self.assertEqual(hosts[0].ip, "192.168.1.2")
    
    def test_import_with_missing_fields(self):
        """Test importing hosts with missing required fields."""
        data = [
            ["Server1", "192.168.1.1", "Location 1", "Group1"],  # Valid
            ["", "192.168.1.2", "Location 2", "Group2"],         # Missing name
            ["Server3", "", "Location 3", "Group1"],             # Missing IP
        ]
        filepath = self.create_test_excel("import.xlsx", data)
        
        existing_ips = set()
        hosts, skipped, errors = ExcelService.import_hosts(filepath, existing_ips)
        
        # Should import only valid host
        self.assertEqual(len(hosts), 1)
        self.assertEqual(skipped, 2)
        self.assertEqual(hosts[0].name, "Server1")
    
    def test_import_with_invalid_ip(self):
        """Test importing hosts - validation happens later."""
        data = [
            ["Server1", "192.168.1.1", "Location 1", "Group1"],
            ["Server2", "example.com", "Location 2", "Group2"],  # Hostname instead
        ]
        filepath = self.create_test_excel("import.xlsx", data)
        
        existing_ips = set()
        hosts, skipped, errors = ExcelService.import_hosts(filepath, existing_ips)
        
        # Both should import successfully (validation happens in Host model)
        self.assertEqual(len(hosts), 2)
        self.assertEqual(skipped, 0)
    
    def test_import_with_long_fields(self):
        """Test importing hosts with fields exceeding length limits."""
        long_name = "A" * 100  # Will be truncated to 50
        long_address = "B" * 200  # Will be truncated to 150
        
        data = [
            [long_name, "192.168.1.1", long_address, "Group1"],
        ]
        filepath = self.create_test_excel("import.xlsx", data)
        
        existing_ips = set()
        hosts, skipped, errors = ExcelService.import_hosts(filepath, existing_ips)
        
        # Should import with truncated fields
        self.assertEqual(len(hosts), 1)
        self.assertEqual(len(hosts[0].name), 50)
        self.assertEqual(len(hosts[0].address), 150)
        
        # Should have error messages about truncation
        self.assertGreater(len(errors), 0)
    
    def test_import_nonexistent_file(self):
        """Test importing from non-existent file."""
        filepath = os.path.join(self.temp_dir, "nonexistent.xlsx")
        
        existing_ips = set()
        
        with self.assertRaises(Exception):
            ExcelService.import_hosts(filepath, existing_ips)
    
    def test_export_with_status_formatting(self):
        """Test that export includes status information."""
        host = TestFixtures.create_sample_host(status="OFFLINE")
        filepath = os.path.join(self.temp_dir, "status.xlsx")
        
        ExcelService.export_hosts(filepath, [host])
        
        # Verify file was created
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active
        
        # Find status column
        headers = [cell.value for cell in sheet[1]]
        status_col = headers.index("Статус") + 1
        
        # Check status value
        status_value = sheet.cell(row=2, column=status_col).value
        self.assertEqual(status_value, "Offline")


if __name__ == '__main__':
    unittest.main()
