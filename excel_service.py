from typing import List, Tuple, Dict, Set
import logging
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from pathlib import Path
from datetime import datetime
from models import Host, HostStatus

class ExcelService:
    @staticmethod
    def import_hosts(file_path: str, existing_ips: Set[str]) -> Tuple[List[Host], int, List[str]]:
        """
        Импорт хостов из Excel.
        Возвращает кортеж (список новых хостов, количество пропущенных, список ошибок).
        """
        workbook = None
        new_hosts = []
        skipped_count = 0
        errors = []

        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            sheet = workbook.active
            
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row or len(row) < 4:
                    continue

                try:
                    name = str(row[0]).strip() if row[0] else ""
                    ip = str(row[1]).strip() if row[1] else ""
                    address = str(row[2]).strip() if row[2] else ""
                    group = str(row[3]).strip() if row[3] else "Без группы"

                    # Применяем ограничения символов (Slicing)
                    if len(name) > 50:
                        name = name[:50]
                        errors.append(f"Строка {row_idx}: название обрезано до 50 символов")
                    
                    if len(ip) > 100:
                        ip = ip[:100]

                    if len(address) > 150:
                        address = address[:150]
                        errors.append(f"Строка {row_idx}: адрес обрезан до 150 символов")

                    if len(group) > 50:
                        group = group[:50]

                    # Проверка обязательных полей
                    if not name or not ip:
                        skipped_count += 1
                        errors.append(f"Строка {row_idx}: пропущена (нет названия или IP)")
                        continue

                    # Проверка на дубликаты
                    if ip in existing_ips:
                        skipped_count += 1
                        errors.append(f"Строка {row_idx}: узел с IP {ip} уже существует")
                        continue

                    # Создаем новый хост
                    try:
                        new_host = Host(
                            name=name,
                            ip=ip,
                            address=address,
                            group=group,
                            notifications_enabled=True
                        )
                    except (ValueError, TypeError) as e:
                        skipped_count += 1
                        errors.append(f"Строка {row_idx}: ошибка создания узла - {e}")
                        continue

                    new_hosts.append(new_host)
                    existing_ips.add(ip)

                except Exception as e:
                    skipped_count += 1
                    errors.append(f"Строка {row_idx}: {str(e)}")

        except Exception as e:
            raise e
        finally:
            if workbook:
                try:
                    workbook.close()
                except Exception as e:
                    logging.debug(f"Ошибка при закрытии workbook: {e}")
        
        return new_hosts, skipped_count, errors

    @staticmethod
    def export_hosts(file_path: str, hosts: List[Host]) -> None:
        """Экспорт хостов в Excel"""
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Network Hosts"

            # Заголовки
            headers = ["Название", "IP адрес", "Адрес", "Группа", "Статус",
                       "Последний ответ", "Offline с", "Уведомления"]
            sheet.append(headers)

            # Стилизация заголовков
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for cell in sheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

            # Данные узлов
            for host in hosts:
                status_name = HostStatus[host.status].title
                last_seen = ""
                if host.last_seen:
                    try:
                        dt = datetime.fromisoformat(host.last_seen)
                        last_seen = dt.strftime("%Y-%m-%d %H:%M:%S")
                    except (ValueError, TypeError) as e:
                        logging.debug(f"Ошибка парсинга last_seen для {host.name}: {e}")

                offline_since = ""
                if host.offline_since:
                    try:
                        dt = datetime.fromisoformat(host.offline_since)
                        offline_since = dt.strftime("%Y-%m-%d %H:%M:%S")
                    except (ValueError, TypeError) as e:
                        logging.debug(f"Ошибка парсинга offline_since для {host.name}: {e}")

                notifications = "Вкл" if host.notifications_enabled else "Выкл"

                row_data = [
                    host.name,
                    host.ip,
                    host.address,
                    host.group,
                    status_name,
                    last_seen,
                    offline_since,
                    notifications
                ]
                sheet.append(row_data)

                # Добавляем границы для последней строки
                for col in range(1, 9):
                    sheet.cell(row=sheet.max_row, column=col).border = border

            # Автоподбор ширины колонок
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except (AttributeError, TypeError):
                        pass  # Игнорируем ячейки с некорректными данными
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width

            workbook.save(file_path)

        except Exception as e:
            raise e
